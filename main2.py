import tkinter as tk
from tkinter import ttk
import  openpyxl
import  pandas as pd
import test as t
from openpyxl import load_workbook
from collections import defaultdict
def load_data():
    path = 'C://Users//ADMIN//PycharmProjects//GUI_csv//tkinter-excel-app//LichThiDau.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_value = list(sheet.values)
    #print(list_value)
    for col_name in list_value[0]:
        treeview.heading(col_name,text=col_name)


    for value_tuple in list_value[1:]:
        treeview.insert('',tk.END,values= value_tuple)


def insert_row():
    changdua = name_ChangDua.get()
    ngaydua = ngayDua.get()
    doidua = name_DoiDua.get()
    taydua = boxTayDua.get()

    print((changdua,ngaydua,doidua,taydua))

    path = 'C://Users//ADMIN//PycharmProjects//GUI_csv//tkinter-excel-app//LichThiDau.xlsx'
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [changdua, doidua, taydua, ngaydua]

    sheet.append(row_values)
    workbook.save(path)



    treeview.insert('', tk.END, values=row_values)


def displayDangKyThiDau():
    widgets_frame2.grid(row=0,column=1, padx=20, pady=10)
    widgets_frame2.config(text= name_ChangDua.get())

    ngay_dua_data = t.csvChangDua.loc[t.csvChangDua['Race'] == name_ChangDua.get(), 'Time']
    ngay_dua_data = ngay_dua_data.to_string(index=False)
    ngayDua.delete(0, 'end')
    ngayDua.insert(0,ngay_dua_data) #########################################


def comTayDua():
    tay_dua_data = t.csvDoiDua.loc[t.csvDoiDua['RacingTeam'] == name_DoiDua.get(), 'Racer']
    tay_dua_data = tay_dua_data.to_string(index=False)
    tay_dua_data = tay_dua_data.split(',')
    tay_dua_data = [string.strip() for string in tay_dua_data]
    label1.config(text=tay_dua_data[0])
    label2.config(text=tay_dua_data[1])


    # boxTayDua = ttk.Combobox(widgets_frame2, values=tay_dua_data)
    # boxTayDua.current(0)
    # boxTayDua.grid(row=2, column=2, padx=5, pady=5, sticky='ew')
    boxTayDua.set(tay_dua_data[0])
    boxTayDua.config(values=tay_dua_data)


def fixTayDua():
    aaa = boxTayDua.get()
#csvChangDua = pd.read_csv('C://Users//ADMIN//PycharmProjects//GUI_csv//tkinter-excel-app//ChangDua.csv')
listChangDua = list(t.csvChangDua['Race'])
listDoiDua = list(t.csvDoiDua['RacingTeam'])

root = tk.Tk()
#root.geometry('1200x741+100+25')
style = ttk.Style(root)
root.tk.call('source','forest-light.tcl')
root.tk.call('source','forest-dark.tcl')
style.theme_use('forest-light')

frame = ttk.Frame(root)
frame.pack()

############# phan 1 ###############
widgets_frame = ttk.LabelFrame(frame,text='Chọn chặng đua')
widgets_frame.grid(row=0,column=0, padx=20, pady=10)

name_ChangDua = ttk.Combobox(widgets_frame, values=listChangDua) ####################################
name_ChangDua.current(0)
name_ChangDua.grid(row=0, column=0,padx=5, pady=5, sticky='ew')

ngay_dua_data = t.csvChangDua.loc[t.csvChangDua['Race'] == name_ChangDua.get(), 'Time']
ngay_dua_data = ngay_dua_data.to_string(index=False)
ngayDua = ttk.Entry(widgets_frame)  ##########################################
ngayDua.insert(0,ngay_dua_data)
ngayDua.grid(row=1, column=0,padx=5, pady=5, sticky='ew')

button = ttk.Button(widgets_frame,text='Xác nhận', command=displayDangKyThiDau)
button.grid(row=2, column=0, padx=5, pady=5,sticky='ew')

########## phan 2 #####################
widgets_frame2 = ttk.LabelFrame(frame,text='            ')
#widgets_frame2.grid(row=0,column=1, padx=20, pady=10)

labelDoiDua = ttk.Label(widgets_frame2,text='Đội đua')
labelDoiDua.grid(row=0,column=0, sticky='ew',padx=5, pady=5)
name_DoiDua = ttk.Combobox(widgets_frame2, values=listDoiDua) ########################
name_DoiDua.current(0)
name_DoiDua.grid(row=0, column=1,padx=5, pady=5, sticky='ew')
butDoiDua = ttk.Button(widgets_frame2,text='Xác nhận', command=comTayDua) ##############
butDoiDua.grid(row=0,column=2, sticky='ew',padx=5, pady=5)

labelTayDua = ttk.LabelFrame(widgets_frame2,text='Tay đua')
labelTayDua.grid(row=2,column=0,columnspan=2, padx=20, pady=10)

tay_dua_data = t.csvDoiDua.loc[t.csvDoiDua['RacingTeam'] == name_DoiDua.get(), 'Racer']
tay_dua_data = tay_dua_data.to_string(index=False)
tay_dua_data = tay_dua_data.split(',')
tay_dua_data = [string.strip() for string in tay_dua_data]

boxTayDua = ttk.Combobox(widgets_frame2, values=tay_dua_data) ##########################
boxTayDua.current(0)
boxTayDua.grid(row=2, column=2, padx=5, pady=5, sticky='ew')



label1 = ttk.Label(labelTayDua,text='    ')
label1.grid(row=0,column=0, padx=20, pady=10)
label2 = ttk.Label(labelTayDua,text='    ')
label2.grid(row=0,column=1, padx=20, pady=10)

butDangKy = ttk.Button(widgets_frame2,text='Đăng ký',command=insert_row)
butDangKy.grid(row=4, column=0,columnspan =3, padx=20, pady=20, sticky='ew')


########## Bang ##############
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=3, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side='right',fill='y')


cols = ("ChangDua", "DoiDua", "TayDua", "NgayDua")
treeview= ttk.Treeview(treeFrame, show='headings',
                       yscrollcommand=treeScroll.set,columns=cols,height=13)
treeview.column('ChangDua',width=100)
treeview.column('DoiDua',width=100)
treeview.column('TayDua',width=100)
treeview.column('NgayDua',width=100)

treeview.pack()
treeScroll.config(command=treeview.yview())
load_data()
root.mainloop()